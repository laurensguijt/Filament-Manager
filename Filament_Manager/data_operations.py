import os
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from Filament_Manager.models import FilamentData, PrintLogEntry

# Ensure Excel file exists
script_directory = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
excel_file = os.path.join(script_directory, "filament_data.xlsx")


def init_excel():
    """Initialize the Excel file if it doesn't exist"""
    if not os.path.exists(excel_file):
        workbook = openpyxl.Workbook()
        
        # Configure Filaments sheet
        filaments_sheet = workbook.active
        filaments_sheet.title = "Filament_Data"
        headers = ["id", "color", "variant", "supplier", "date_opened", "weight_g", "hex_color", "empty_spool_weight", "description"]
        filaments_sheet.append(headers)
        
        # Set column widths for Filaments sheet
        for col in range(1, len(headers) + 1):
            filaments_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Create and configure Print Log sheet
        print_log_sheet = workbook.create_sheet("Print_Log")
        log_headers = ["timestamp", "print_name", "filament_code", "material", "variant", "used_weight", "remaining_weight"]
        print_log_sheet.append(log_headers)
        
        # Set column widths for Print Log sheet
        for col in range(1, len(log_headers) + 1):
            print_log_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
        workbook.save(excel_file)


def read_excel_data():
    """Read filament data from Excel file"""
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Filament_Data"]
    data = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Skip empty rows
            try:
                filament = FilamentData.from_row(row)
                data.append(filament)
            except Exception as e:
                continue
    
    return data


def write_excel_data(data):
    """Write filament data to Excel file"""
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Filament_Data"]
    
    # Clear existing data (except headers)
    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None
    
    # Write new data
    for row_idx, filament in enumerate(data, start=2):
        row_data = filament.to_row()
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # If this is the hex_color column (column 7), set the cell background
            if col_idx == 7 and isinstance(value, str) and value.startswith('#'):
                cell.fill = PatternFill(start_color=value.replace('#', ''), 
                                      end_color=value.replace('#', ''),
                                      fill_type='solid')
    
    try:
        workbook.save(excel_file)
    except Exception as e:
        raise


def add_print_log_entry(timestamp, print_name, filament_code, material, variant, used_weight, remaining_weight):
    """Add a new entry to the print log"""
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Print_Log"]
    
    # Create PrintLogEntry
    entry = PrintLogEntry(
        timestamp=timestamp,
        print_name=print_name,
        filament_code=filament_code,
        material=material,
        variant=variant,
        used_weight=used_weight,
        remaining_weight=remaining_weight
    )
    
    # Add new row
    sheet.append(entry.to_row())
    workbook.save(excel_file)


def read_print_log():
    """Read all entries from the print log"""
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Print_Log"]
    log_entries = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Skip empty rows
            log_entries.append(PrintLogEntry.from_row(row))
    return log_entries


def get_next_code(data):
    """Generate the next available filament code"""
    if not data:
        return "F001"  # First filament
    
    # Get existing codes and find the highest number
    existing_codes = [filament.code for filament in data if filament.code.startswith('F')]
    if not existing_codes:
        return "F001"
    
    # Get the highest number
    highest = max(int(code[1:]) for code in existing_codes if code[1:].isdigit())
    # Generate next code
    return f"F{(highest + 1):03d}" 